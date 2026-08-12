#!/usr/bin/env python3
"""
データエクスポートモジュール

処理されたデータと解析結果をExcelファイルにエクスポートする機能を提供します。
重力レベルデータとグラフ、G-quality解析結果を保存します。
"""

from __future__ import annotations

import contextlib
import os
import shutil
import tempfile
from collections.abc import Callable
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from core.exceptions import ExportError
from core.logger import get_logger, log_exception
from core.paths import ensure_graphs_dir, ensure_results_dir

# モジュール用のロガーを初期化
logger = get_logger("export")

ConfirmHandler = Callable[[Path], bool]
NotifyHandler = Callable[[str], None]


def _default_confirm_overwrite(path: Path) -> bool:
    logger.warning("上書き確認のハンドラが指定されていないため自動的に上書きします: %s", path)
    return True


def _default_notify_warning(message: str) -> None:
    logger.warning(message)


def _default_notify_info(message: str) -> None:
    logger.info(message)


# 共通時間軸の上限。1サンプル/秒でも約11日分あり、設定ミスで数十GBの
# DataFrameを作ってプロセスが落ちるのを防ぐ。
MAX_UNIFIED_SAMPLES = 20_000_000

G_QUALITY_SHEET_NAME = "G-quality Analysis"


def _save_workbook_atomically(workbook, target: Path) -> None:
    """一時ファイル経由でワークブックを保存する

    openpyxl は保存先を切り詰めてから書き込むため、途中で失敗すると壊れた
    ファイルだけが残る。同じディレクトリの一時ファイルへ書き、成功した場合に
    限って os.replace で差し替える。
    """
    target = Path(target)
    fd, temp_name = tempfile.mkstemp(prefix=f".{target.stem}_", suffix=target.suffix, dir=str(target.parent))
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        workbook.save(temp_path)
        os.replace(temp_path, target)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _read_sheet_rows(workbook_path: Path, sheet_name: str) -> list[tuple]:
    """既存ワークブックから指定シートの内容を読み出す（無ければ空リスト）"""
    if not workbook_path.exists():
        return []
    workbook = None
    try:
        workbook = load_workbook(workbook_path, read_only=True)
        if sheet_name not in workbook.sheetnames:
            return []
        return [row for row in workbook[sheet_name].iter_rows(values_only=True) if any(v is not None for v in row)]
    except Exception as e:
        logger.warning(f"既存の '{sheet_name}' シートを読み込めませんでした: {e}")
        return []
    finally:
        if workbook is not None:
            with contextlib.suppress(Exception):  # pragma: no cover - read_only handle cleanup
                workbook.close()


def _build_unified_time_axis(start_time: float, end_time: float, sampling_rate: float) -> np.ndarray:
    """開始・終了時刻から共通時間軸を作る

    np.arange(start, end + step, step) は浮動小数点誤差で end を1サンプル
    超える要素を混ぜることがあり、その点は補間で端点にクランプされて実測に
    ない重複行になる。サンプル数を先に決めて生成することで境界を確定させる。
    """
    if end_time < start_time:
        start_time, end_time = end_time, start_time

    span = float(end_time) - float(start_time)
    sample_count = int(round(span * sampling_rate)) + 1
    if sample_count > MAX_UNIFIED_SAMPLES:
        raise ExportError(
            f"共通時間軸のサンプル数が上限を超えました ({sample_count} > {MAX_UNIFIED_SAMPLES})。"
            "サンプリングレートまたはトリミング範囲の設定を確認してください。"
        )
    sample_count = max(sample_count, 1)
    return float(start_time) + np.arange(sample_count, dtype=np.float64) / sampling_rate


def _resample_to_axis(unified_time: np.ndarray, times: pd.Series, values: pd.Series) -> np.ndarray:
    """センサー系列を共通時間軸へ線形補間する

    np.interp は x が昇順であることを前提とし、そうでない場合は黙って
    誤った値を返すため、事前に時間順へ並べ替える。実測範囲の外側は
    端点にクランプされた偽の値ではなく NaN（Excelでは空欄）にする。
    """
    time_values = np.asarray(times, dtype=np.float64)
    data_values = np.asarray(values, dtype=np.float64)
    if time_values.size != data_values.size:
        raise ExportError(
            f"時間と測定値の長さが一致しません: time={time_values.size}, values={data_values.size}",
        )

    usable = np.isfinite(time_values)
    if not usable.any():
        return np.full(unified_time.shape, np.nan)

    time_values = time_values[usable]
    data_values = data_values[usable]

    order = np.argsort(time_values, kind="stable")
    if not np.all(order == np.arange(time_values.size)):
        logger.warning("時間列が昇順ではないため、補間前に時間順へ並べ替えます。")
        time_values = time_values[order]
        data_values = data_values[order]

    resampled = np.interp(unified_time, time_values, data_values)
    outside = (unified_time < time_values[0]) | (unified_time > time_values[-1])
    resampled[outside] = np.nan
    return resampled


def create_output_directories(csv_dir: str | None = None) -> tuple[Path, Path]:
    """
    出力用ディレクトリ構造を作成する

    CSVファイルがあるディレクトリに `results_AAT/graphs` ディレクトリを作成します。
    csv_dir が指定されていない場合は、プロジェクトルートにフォールバックします。

    Args:
        csv_dir (str, optional): CSVファイルのディレクトリパス。指定されていればそのディレクトリに作成

    Returns:
        tuple: 作成した結果ディレクトリとグラフディレクトリのパス
    """
    logger.debug("create_output_directories called with csv_dir: %s", csv_dir)

    results_dir = ensure_results_dir(csv_dir)
    graphs_dir = ensure_graphs_dir(csv_dir)

    logger.debug("Created directories: results=%s, graphs=%s", results_dir, graphs_dir)

    return results_dir, graphs_dir


def export_data(
    time: pd.Series,
    adjusted_time: pd.Series,
    gravity_level_inner_capsule: pd.Series,
    gravity_level_drag_shield: pd.Series,
    file_path: str,
    min_mean_inner_capsule: float | None,
    min_time_inner_capsule: float | None,
    min_std_inner_capsule: float | None,
    min_mean_drag_shield: float | None,
    min_time_drag_shield: float | None,
    min_std_drag_shield: float | None,
    graph_path: str | None,
    filtered_time: pd.Series,  # フィルタリング済みの時間データを追加  # noqa: ARG001
    filtered_adjusted_time: pd.Series,  # フィルタリング済みの調整時間データを追加  # noqa: ARG001
    config: dict[str, Any] | None = None,  # 設定パラメータを追加
    raw_data: pd.DataFrame | None = None,  # 元のCSVデータを追加
    *,
    confirm_overwrite: ConfirmHandler | None = None,
    notify_warning: NotifyHandler | None = None,
    notify_info: NotifyHandler | None = None,
) -> str:
    """
    処理されたデータとグラフをExcelにエクスポートする

    重力レベルデータと統計情報を含むExcelファイルを作成します。
    グラフは別途グラフディレクトリに保存します。
    既存のファイルが存在する場合は、上書きするかどうかを確認します。

    Args:
        time (pandas.Series): 時間データ
        adjusted_time (pandas.Series): 調整された時間データ
        gravity_level_inner_capsule (pandas.Series): Inner Capsuleの重力レベル
        gravity_level_drag_shield (pandas.Series): Drag Shieldの重力レベル
        file_path (str): 元のCSVファイルのパス
        min_mean_inner_capsule (float): Inner Capsuleの最小標準偏差ウィンドウの平均値
        min_time_inner_capsule (float): Inner Capsuleの最小標準偏差ウィンドウの開始時間
        min_std_inner_capsule (float): Inner Capsuleの最小標準偏差値
        min_mean_drag_shield (float): Drag Shieldの最小標準偏差ウィンドウの平均値
        min_time_drag_shield (float): Drag Shieldの最小標準偏差ウィンドウの開始時間
        min_std_drag_shield (float): Drag Shieldの最小標準偏差値
        graph_path (str): 保存されたグラフの画像ファイルパス
        filtered_time (pandas.Series): フィルタリングされた時間データ
        filtered_adjusted_time (pandas.Series): フィルタリングされた調整時間データ
        config (dict, optional): 設定パラメータ。指定されない場合はデフォルト値を使用
        raw_data (pandas.DataFrame, optional): 元のCSVデータ。加速度データ出力用

    Returns:
        str: 出力されたExcelファイルのパス

    Raises:
        ExportError: データのエクスポート中にエラーが発生した場合
    """
    confirm_overwrite = confirm_overwrite or _default_confirm_overwrite
    notify_warning = notify_warning or _default_notify_warning
    notify_info = notify_info or _default_notify_info

    # CSVファイルのディレクトリとファイル名を取得
    file_path_obj = Path(file_path)
    csv_dir = str(file_path_obj.parent)
    base_name = file_path_obj.stem

    # 出力ディレクトリ構造を作成
    results_dir, graphs_dir = create_output_directories(csv_dir)

    # 出力ファイルパスの設定（シンプルな名前を使用）
    output_file_path = results_dir / f"{base_name}.xlsx"
    output_stem = base_name

    # 既存ファイルの確認。グラフのコピーより先に決めることで、上書きを断った
    # ときに前回のグラフ画像を壊さないようにする。
    if output_file_path.exists() and not confirm_overwrite(output_file_path):
        # 新しいファイル名を生成（連番を付加）
        counter = 1
        while (results_dir / f"{base_name}_{counter}.xlsx").exists():
            counter += 1
        output_stem = f"{base_name}_{counter}"
        output_file_path = results_dir / f"{output_stem}.xlsx"
        notify_info(f"ファイル名を変更して保存します: {output_file_path.name}")

    # グラフファイルのパスはワークブックと同じ語幹に揃える
    new_graph_path = graphs_dir / f"{output_stem}_gl.png"

    # 既存グラフが元のパスに存在する場合、新しいパスにコピー
    if graph_path is not None:
        graph_path_obj = Path(graph_path)
        if graph_path_obj.exists() and graph_path_obj.resolve() != new_graph_path.resolve():
            shutil.copy2(graph_path, new_graph_path)

    try:
        # 共通の時間軸を作成
        time_ranges = []
        if time is not None and not time.empty:
            time_ranges.append((time.min(), time.max()))
        if adjusted_time is not None and not adjusted_time.empty:
            time_ranges.append((adjusted_time.min(), adjusted_time.max()))

        if not time_ranges:
            raise ExportError("エクスポート可能な時間データがありません。")

        # 両センサーの和集合を使う。交差範囲に切ると、長い側のセンサーで実測した
        # 区間が出力から消え、Statisticsシート（各センサーの全区間で算出）と
        # 対応が取れなくなる。各センサーの実測範囲外は補間せず空欄にする。
        start_time = min(range_[0] for range_ in time_ranges)
        end_time = max(range_[1] for range_ in time_ranges)

        # configが指定されていない場合のデフォルト値
        if config is None:
            config = {}

        # 時間間隔を計算（サンプリングレートに基づく）
        sampling_rate = config.get("sampling_rate", 1000)  # 設定からサンプリングレートを取得、デフォルトは1000Hz
        if not sampling_rate or sampling_rate <= 0:
            raise ValueError("サンプリングレートは正の数でなければなりません。")

        unified_time = _build_unified_time_axis(start_time, end_time, float(sampling_rate))

        # データフレームの作成（統一された時間軸）
        export_columns = {"Time (s)": unified_time}
        if (
            time is not None
            and not time.empty
            and gravity_level_inner_capsule is not None
            and not gravity_level_inner_capsule.empty
        ):
            export_columns["Gravity Level (Inner Capsule) (G)"] = _resample_to_axis(
                unified_time, time, gravity_level_inner_capsule
            )
        if (
            adjusted_time is not None
            and not adjusted_time.empty
            and gravity_level_drag_shield is not None
            and not gravity_level_drag_shield.empty
        ):
            export_columns["Gravity Level (Drag Shield) (G)"] = _resample_to_axis(
                unified_time, adjusted_time, gravity_level_drag_shield
            )
        export_data = pd.DataFrame(export_columns)

        # 統計情報のデータフレームを作成
        stats_df = pd.DataFrame(
            {
                "Statistic": [
                    "Inner Capsule: Mean Gravity Level of the interval with the smallest standard deviation(G)",
                    "Inner Capsule: Time at smallest Standard Deviation(s)",
                    "Inner Capsule: smallest Standard Deviation(G)",
                    "Drag Shield: Mean Gravity Level of the interval with the smallest standard deviation(G)",
                    "Drag Shield: Time at smallest Standard Deviation(s)",
                    "Drag Shield: smallest Standard Deviation(G)",
                ],
                "Value": [
                    min_mean_inner_capsule,
                    min_time_inner_capsule,
                    min_std_inner_capsule,
                    min_mean_drag_shield,
                    min_time_drag_shield,
                    min_std_drag_shield,
                ],
            }
        )

        # トリミング範囲の加速度データを準備
        acceleration_data = None
        if raw_data is not None:
            # 元のCSVデータが提供されている場合
            try:
                # 設定から列名を取得
                time_column = config.get("time_column")
                acceleration_inner_column = config.get("acceleration_column_inner_capsule")
                acceleration_drag_column = config.get("acceleration_column_drag_shield")
                use_inner = config.get("use_inner_acceleration", True)
                use_drag = config.get("use_drag_acceleration", True)

                logger.info(
                    "加速度データの処理開始: "
                    f"時間列={time_column}, 内カプセル加速度列={acceleration_inner_column}, 外カプセル加速度列={acceleration_drag_column}, "
                    f"_inner使用={use_inner}, _drag使用={use_drag}"
                )
                logger.debug(f"元データの列: {raw_data.columns.tolist()}")

                if time_column is None:
                    notify_warning("加速度データの時間列が設定されていないため、エクスポートをスキップします。")
                    raw_columns_valid = False
                else:
                    raw_columns_valid = True
                    missing_cols = []
                    if time_column not in raw_data.columns:
                        raw_columns_valid = False
                        missing_cols.append(f"時間列({time_column})")
                    if use_inner and acceleration_inner_column and acceleration_inner_column not in raw_data.columns:
                        raw_columns_valid = False
                        missing_cols.append(f"内カプセル加速度列({acceleration_inner_column})")
                    if use_drag and acceleration_drag_column and acceleration_drag_column not in raw_data.columns:
                        raw_columns_valid = False
                        missing_cols.append(f"外カプセル加速度列({acceleration_drag_column})")

                    if not raw_columns_valid and missing_cols:
                        logger.warning(f"必要な列が見つかりません: {missing_cols}")
                        notify_warning(
                            "必要な列が見つかりません: "
                            f"\n{', '.join(missing_cols)}\n\n"
                            "加速度データがシートに追加されません。\n"
                            "CSVファイルを選択する際、正しい列を選んでください。"
                        )

                if raw_columns_valid:
                    try:
                        orig_time_data = raw_data[time_column].values.astype(float)
                        acceleration_columns: dict[str, np.ndarray] = {}

                        if use_inner and acceleration_inner_column:
                            acceleration_columns["inner"] = raw_data[acceleration_inner_column].values.astype(float)
                        if use_drag and acceleration_drag_column:
                            acceleration_columns["drag"] = raw_data[acceleration_drag_column].values.astype(float)

                        if not acceleration_columns:
                            logger.info(
                                "有効な加速度列が選択されていないため、加速度データのエクスポートをスキップします"
                            )
                        else:
                            # data_processor.load_and_process_data と同じ同期点ロジックで
                            # 各センサーの時間軸を調整する（Gravity Levelシートとの時間軸整合のため）
                            acc_thresh = config.get("acceleration_threshold", 1.0)

                            def _first_sync_index(values: np.ndarray) -> int | None:
                                sync_mask = np.abs(values) < acc_thresh
                                return int(np.where(sync_mask)[0][0]) if sync_mask.any() else None

                            drag_sync_idx = (
                                _first_sync_index(acceleration_columns["drag"])
                                if "drag" in acceleration_columns
                                else None
                            )
                            inner_sync_idx = (
                                _first_sync_index(acceleration_columns["inner"])
                                if "inner" in acceleration_columns
                                else None
                            )
                            # Innerの同期点が見つからない場合はDrag Shieldの同期点を流用し、
                            # どちらも見つからない場合は先頭サンプルを使用する（data_processorと同じ挙動）
                            if inner_sync_idx is None:
                                inner_sync_idx = drag_sync_idx if drag_sync_idx is not None else 0
                            if drag_sync_idx is None:
                                drag_sync_idx = 0

                            accel_frame = {"Time (s)": unified_time}
                            if "inner" in acceleration_columns:
                                inner_adjusted_time = orig_time_data - orig_time_data[inner_sync_idx]
                                accel_frame["Acceleration (Inner Capsule) (m/s²)"] = _resample_to_axis(
                                    unified_time, inner_adjusted_time, acceleration_columns["inner"]
                                )
                            if "drag" in acceleration_columns:
                                drag_adjusted_time = orig_time_data - orig_time_data[drag_sync_idx]
                                accel_frame["Acceleration (Drag Shield) (m/s²)"] = _resample_to_axis(
                                    unified_time, drag_adjusted_time, acceleration_columns["drag"]
                                )

                            acceleration_data = pd.DataFrame(accel_frame)
                            logger.info(f"共通時間軸で加速度データを作成: {len(acceleration_data)}行")

                    except Exception as e:
                        log_exception(e, "加速度データのエクスポート中にエラーが発生しました")
                        notify_warning(f"加速度データの保存中にエラーが発生しました: {e}")
                        acceleration_data = None
            except Exception as e:
                log_exception(e, "加速度データの準備中にエラーが発生しました")
                acceleration_data = None

        # ExcelWriterは書き込みモードでファイル全体を置き換えるため、既存の
        # G-quality Analysisシートを先に退避して書き戻す。そうしないと
        # G-quality自動計算が無効な状態で再処理したときに前回の解析結果が消える。
        preserved_g_quality = _read_sheet_rows(output_file_path, G_QUALITY_SHEET_NAME)

        # 出力先へ直接書くと、書き込み中の失敗で前回の結果ファイルまで
        # 読めない状態で失われる。一時ファイルへ書き切ってから置き換える。
        fd, temp_name = tempfile.mkstemp(
            prefix=f".{output_file_path.stem}_", suffix=".xlsx", dir=str(output_file_path.parent)
        )
        os.close(fd)
        temp_workbook_path = Path(temp_name)
        try:
            with pd.ExcelWriter(temp_workbook_path, engine="openpyxl") as writer:
                export_data.to_excel(writer, sheet_name="Gravity Level Data", index=False)
                stats_df.to_excel(writer, sheet_name="Gravity Level Statistics", index=False)
                if acceleration_data is not None:
                    acceleration_data.to_excel(writer, sheet_name="Acceleration Data", index=False)
                    logger.info(f"加速度データをシートに追加しました: {len(acceleration_data)}行")
                else:
                    logger.warning("加速度データが作成されなかったため、シートに追加されません")
                if preserved_g_quality:
                    sheet = writer.book.create_sheet(title=G_QUALITY_SHEET_NAME)
                    for row in preserved_g_quality:
                        sheet.append(list(row))
                    logger.info(f"既存のG-quality解析シートを引き継ぎました: {len(preserved_g_quality) - 1}行")
            os.replace(temp_workbook_path, output_file_path)
        except Exception:
            temp_workbook_path.unlink(missing_ok=True)
            raise

        graph_exists = graph_path is not None and Path(graph_path).exists()
        graph_display_target = new_graph_path if graph_exists else new_graph_path.parent
        graphs_message = (
            f"- グラフ画像: {graph_display_target}" if graph_exists else f"- グラフ出力フォルダ: {graph_display_target}"
        )
        message = f"保存が完了しました。\n- Gravity Levelデータ: {output_file_path}\n{graphs_message}"
        notify_info(message)

        return str(output_file_path)
    except PermissionError as e:
        error_msg = f"{output_file_path} に書き込みできません。権限を確認してください。"
        logger.error(error_msg)
        raise ExportError(error_msg, file_path=str(output_file_path)) from e
    except Exception as e:
        error_msg = f"データの保存中にエラーが発生しました: {e}"
        log_exception(e, error_msg)
        raise ExportError(error_msg, file_path=str(output_file_path)) from e


def export_g_quality_data(g_quality_data, original_file_path, g_quality_graph_path=None, workbook_path=None):
    """
    G-quality解析の結果をエクスポートする

    異なるウィンドウサイズでのG-quality評価結果をExcelファイルに追加または新規作成します。
    既存のExcelファイルがある場合は、G-quality Analysis シートを更新します。
    また、G-qualityグラフが提供されている場合は、指定されたディレクトリにコピーします。

    Args:
        g_quality_data (list): G-quality解析の結果データ
            各要素は (window_size, min_time_inner_capsule, min_mean_inner_capsule, min_std_inner_capsule,
                      min_time_drag_shield, min_mean_drag_shield, min_std_drag_shield) の形式のタプル
        original_file_path (str): 元のCSVファイルのパス
        g_quality_graph_path (str, optional): G-qualityグラフの画像ファイルパス。指定された場合はグラフをコピーします。
        workbook_path (str | Path, optional): 書き込み先のExcelファイル。``export_data`` が
            上書きを避けて連番付きの名前で保存した場合に、そのパスを渡すことで
            重力レベルデータとG-quality結果が同じワークブックに揃う。省略時は
            ``<CSV名>.xlsx`` を使用する。

    Returns:
        str or None: 出力されたExcelファイルのパス、または失敗した場合はNone

    Raises:
        ExportError: データのエクスポート中にエラーが発生した場合
    """
    # CSVファイルのディレクトリとファイル名を取得
    file_path_obj = Path(original_file_path)
    csv_dir = str(file_path_obj.parent)
    base_name = file_path_obj.stem

    # 出力ディレクトリ構造を作成
    results_dir, graphs_dir = create_output_directories(csv_dir)

    # 出力ファイルパスの設定。グラフ名もワークブックの語幹に揃える。
    if workbook_path is not None:
        output_file_path = Path(workbook_path)
        graph_stem = output_file_path.stem
    else:
        output_file_path = results_dir / f"{base_name}.xlsx"
        graph_stem = base_name

    # G-qualityグラフの処理
    if g_quality_graph_path and Path(g_quality_graph_path).exists():
        # グラフファイルの新しいパスを設定（短い名前を使用）
        new_graph_path = graphs_dir / f"{graph_stem}_gq.png"

        # 同じファイルでない場合のみコピーを実行

        try:
            # パスを正規化して比較
            source_path = Path(g_quality_graph_path).resolve()
            dest_path = new_graph_path.resolve()

            if source_path != dest_path:
                shutil.copy2(g_quality_graph_path, new_graph_path)
                logger.info(f"G-qualityグラフを保存しました: {g_quality_graph_path} -> {new_graph_path}")
            else:
                logger.debug(f"G-qualityグラフは既に正しい場所にあります: {new_graph_path}")
        except Exception as e:
            logger.warning(f"G-qualityグラフの保存中にエラーが発生しました: {e}")

    # データフレームの作成
    df = pd.DataFrame(
        g_quality_data,
        columns=[
            "Window Size (s)",
            "Inner Capsule: Time at smallest Standard Deviation(s)",
            "Inner Capsule: Mean Gravity Level of the interval with the smallest standard deviation(G)",
            "Inner Capsule: smallest Standard Deviation(G)",
            "Drag Shield: Time at smallest Standard Deviation(s)",
            "Drag Shield: Mean Gravity Level of the interval with the smallest standard deviation(G)",
            "Drag Shield: smallest Standard Deviation(G)",
        ],
    )

    try:
        # 既存のExcelファイルを読み込むか、新規作成
        try:
            workbook = load_workbook(output_file_path)
        except FileNotFoundError:
            workbook = Workbook()
            # デフォルトのシートを削除（後で必要なシートを追加する）
            sheet_to_remove = workbook.active
            if sheet_to_remove is not None:
                workbook.remove(sheet_to_remove)

        # G-quality Analysis シートを作成または更新
        if G_QUALITY_SHEET_NAME in workbook.sheetnames:
            del workbook[G_QUALITY_SHEET_NAME]
        sheet = workbook.create_sheet(title=G_QUALITY_SHEET_NAME)

        # データをシートに書き込む（1行目から開始）
        for row in dataframe_to_rows(df, index=False, header=True):
            sheet.append(row)

        # 一時ファイルへ書いてから置き換える。保存中に失敗しても、既存の
        # ワークブックが半端な状態で残らないようにする。
        _save_workbook_atomically(workbook, output_file_path)
        return output_file_path
    except PermissionError as e:
        raise ExportError(
            f"{output_file_path} に書き込みできません。ファイルが開かれている可能性があります。",
            file_path=str(output_file_path),
        ) from e
    except Exception as e:
        raise ExportError(f"データの保存中にエラーが発生しました: {e}", file_path=str(output_file_path)) from e
